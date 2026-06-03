from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, Http404
from django.db.models import Sum
from django_filters.rest_framework import DjangoFilterBackend
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Employee, Project, TimesheetEntry, ProjectHours, EmployeeProject
from .serializers import (
    EmployeeSerializer, ProjectSerializer,
    TimesheetEntrySerializer, ProjectHoursSerializer, EmployeeProjectSerializer
)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def superuser_required(view_func):
    """Redirect non-superusers to their own timesheet."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('my_timesheet')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─── Auth Views ──────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('dashboard')
        return redirect('my_timesheet')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('dashboard' if user.is_superuser else 'my_timesheet')
        else:
            error = 'Invalid username or password.'
    return render(request, 'core/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('login')


# ─── Superuser Views ─────────────────────────────────────────────────────────

@superuser_required
def dashboard(request):
    today = datetime.date.today()
    projects = Project.objects.filter(is_active=True)
    employees = Employee.objects.filter(is_active=True)

    today_entries = TimesheetEntry.objects.filter(date=today).select_related('employee').prefetch_related('project_hours__project')

    week_no = today.isocalendar()[1]
    week_hours = TimesheetEntry.objects.filter(year=today.year, week_number=week_no).aggregate(total=Sum('total_hours'))['total'] or 0

    month_project_hours = (
        ProjectHours.objects
        .filter(entry__year=today.year, entry__month_number=today.month)
        .values('project__code', 'project__name')
        .annotate(total=Sum('hours'))
        .order_by('-total')
    )

    context = {
        'projects': projects,
        'employees': employees,
        'today': today,
        'today_entries': today_entries,
        'week_hours': week_hours,
        'month_project_hours': month_project_hours,
        'total_employees': employees.count(),
        'total_projects': projects.count(),
    }
    return render(request, 'core/dashboard.html', context)


@superuser_required
def projects_view(request):
    projects = Project.objects.all().order_by('-created_at')
    return render(request, 'core/projects.html', {'projects': projects})


@superuser_required
def employees_view(request):
    employees = Employee.objects.all()
    return render(request, 'core/employees.html', {'employees': employees})


@superuser_required
def all_timesheet_view(request):
    """Superuser sees all employees' entries."""
    projects = Project.objects.filter(is_active=True)
    employees = Employee.objects.filter(is_active=True)
    return render(request, 'core/timesheet.html', {
        'projects': projects,
        'employees': employees,
        'is_superuser': True,
    })


@superuser_required
def export_view(request):
    """Export page — pick project + month."""
    projects = Project.objects.filter(is_active=True)
    return render(request, 'core/export.html', {'projects': projects})


@superuser_required
def export_excel(request):
    """Generate and stream the Excel file."""
    project_id = request.GET.get('project')
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not all([project_id, year, month]):
        return redirect('export')

    try:
        project = Project.objects.get(pk=project_id)
        year, month = int(year), int(month)
    except (Project.DoesNotExist, ValueError):
        raise Http404

    # Fetch all entries for this project+month that have hours > 0
    ph_qs = (
        ProjectHours.objects
        .filter(project=project, entry__year=year, entry__month_number=month, hours__gt=0)
        .select_related('entry__employee')
        .order_by('entry__date', 'entry__employee__name')
    )

    month_name = datetime.date(year, month, 1).strftime('%B %Y')

    # ── Build Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{project.code} - {month_name}"

    # Color palette
    HDR_FILL   = PatternFill("solid", fgColor="1A3A5C")
    SUB_FILL   = PatternFill("solid", fgColor="1E2D45")
    ALT_FILL   = PatternFill("solid", fgColor="131929")
    GREEN_FILL = PatternFill("solid", fgColor="1B3A2A")
    thin = Side(style='thin', color='2E3350')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    HDR_FONT  = Font(bold=True, color="E8EAF0", name="Calibri", size=11)
    BODY_FONT = Font(color="C8CADC", name="Calibri", size=10)
    TITLE_FONT = Font(bold=True, color="4F8EF7", name="Calibri", size=14)
    ACCENT_FONT = Font(bold=True, color="3DD68C", name="Calibri", size=10)

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Timesheet — {project.name} ({project.code})"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = PatternFill("solid", fgColor="0F1117")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Period: {month_name}   |   Exported: {datetime.date.today().strftime('%d %b %Y')}"
    ws['A2'].font = Font(color="7B80A0", name="Calibri", size=9, italic=True)
    ws['A2'].fill = PatternFill("solid", fgColor="0F1117")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6  # spacer

    # ── Header row ───────────────────────────────────────────────────────────
    headers = ['Date', 'Week', 'Month', 'Employee', 'Task Description', 'Hours', 'Status']
    col_widths = [14, 8, 8, 22, 60, 10, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    total_hours = 0
    for row_idx, ph in enumerate(ph_qs, 5):
        e = ph.entry
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="161B2C")
        status_map = {'present': 'Present', 'holiday': 'Holiday', 'leave': 'Leave'}

        row_data = [
            e.date.strftime('%d-%m-%Y'),
            e.week_number,
            e.month_number,
            e.employee.name,
            e.task_description or '',
            float(ph.hours),
            status_map.get(e.status, e.status),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = ACCENT_FONT if col == 6 else BODY_FONT
            if col in [1, 2, 3, 6, 7]:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row_idx].height = 30
        total_hours += float(ph.hours)

    # ── Summary row ──────────────────────────────────────────────────────────
    summary_row = len(ph_qs) + 5
    ws.row_dimensions[summary_row].height = 8  # spacer

    total_row = summary_row + 1
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws.cell(total_row, 1, 'TOTAL HOURS').font = Font(bold=True, color="F7C94F", name="Calibri", size=11)
    ws.cell(total_row, 1).fill = PatternFill("solid", fgColor="1A1D27")
    ws.cell(total_row, 1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(total_row, 1).border = border

    tc = ws.cell(total_row, 6, total_hours)
    tc.font = Font(bold=True, color="3DD68C", name="Calibri", size=12)
    tc.fill = GREEN_FILL
    tc.border = border
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[total_row].height = 26

    # Freeze top rows
    ws.freeze_panes = 'A5'

    # ── Stream response ──────────────────────────────────────────────────────
    fname = f"{project.code}_{month_name.replace(' ', '_')}_Timesheet.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# ─── Normal User Views ───────────────────────────────────────────────────────

@login_required
def my_timesheet_view(request):
    """Normal user: sees only their own entries and only their assigned projects."""
    try:
        employee = Employee.objects.get(name__iexact=request.user.get_full_name())
    except Employee.DoesNotExist:
        try:
            employee = Employee.objects.get(name__iexact=request.user.username)
        except Employee.DoesNotExist:
            employee = None

    # Only show projects assigned to this employee by admin
    if employee:
        assigned_ids = EmployeeProject.objects.filter(employee=employee).values_list('project_id', flat=True)
        projects = Project.objects.filter(id__in=assigned_ids, is_active=True)
    else:
        projects = Project.objects.none()

    return render(request, 'core/my_timesheet.html', {
        'projects': projects,
        'employee': employee,
        'user': request.user,
    })


@superuser_required
def role_management_view(request):
    """Superuser assigns projects to employees."""
    employees = Employee.objects.filter(is_active=True).prefetch_related('assigned_projects__project')
    projects = Project.objects.filter(is_active=True)
    return render(request, 'core/role_management.html', {
        'employees': employees,
        'projects': projects,
    })


# ─── DRF API ViewSets ────────────────────────────────────────────────────────

class IsSuperUser(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and request.user.is_superuser


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsSuperUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        project = self.get_object()
        project.is_active = not project.is_active
        project.save()
        return Response({'is_active': project.is_active})


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsSuperUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']


class EmployeeProjectViewSet(viewsets.ModelViewSet):
    """API for assigning/removing projects from employees."""
    queryset = EmployeeProject.objects.all().select_related('employee', 'project')
    serializer_class = EmployeeProjectSerializer
    permission_classes = [IsSuperUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'project']

    @action(detail=False, methods=['post'])
    def assign(self, request):
        """Assign a project to an employee."""
        emp_id = request.data.get('employee')
        proj_id = request.data.get('project')
        if not emp_id or not proj_id:
            return Response({'detail': 'employee and project are required.'}, status=400)
        obj, created = EmployeeProject.objects.get_or_create(
            employee_id=emp_id, project_id=proj_id
        )
        return Response(EmployeeProjectSerializer(obj).data, status=201 if created else 200)

    @action(detail=False, methods=['post'])
    def unassign(self, request):
        """Remove a project assignment from an employee."""
        emp_id = request.data.get('employee')
        proj_id = request.data.get('project')
        deleted, _ = EmployeeProject.objects.filter(employee_id=emp_id, project_id=proj_id).delete()
        if deleted:
            return Response({'detail': 'Unassigned.'}, status=200)
        return Response({'detail': 'Assignment not found.'}, status=404)

    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        """Get all project assignments for a specific employee."""
        emp_id = request.query_params.get('employee')
        if not emp_id:
            return Response({'detail': 'employee param required.'}, status=400)
        qs = self.queryset.filter(employee_id=emp_id)
        return Response(EmployeeProjectSerializer(qs, many=True).data)


class TimesheetEntryViewSet(viewsets.ModelViewSet):
    serializer_class = TimesheetEntrySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['date', 'employee', 'year', 'week_number', 'month_number', 'status']

    def get_queryset(self):
        qs = TimesheetEntry.objects.all().select_related('employee').prefetch_related('project_hours__project')
        # Normal users only see their own entries
        if not self.request.user.is_superuser:
            try:
                emp = Employee.objects.get(name__iexact=self.request.user.get_full_name())
            except Employee.DoesNotExist:
                try:
                    emp = Employee.objects.get(name__iexact=self.request.user.username)
                except Employee.DoesNotExist:
                    return qs.none()
            qs = qs.filter(employee=emp)
        return qs

    @action(detail=False, methods=['get'])
    def summary(self, request):
        if not request.user.is_superuser:
            return Response({'detail': 'Superuser only.'}, status=403)
        year = request.query_params.get('year', datetime.date.today().year)
        month = request.query_params.get('month', datetime.date.today().month)
        data = (
            ProjectHours.objects
            .filter(entry__year=year, entry__month_number=month)
            .values('project__code', 'project__name', 'entry__employee__name')
            .annotate(total_hours=Sum('hours'))
            .order_by('entry__employee__name', 'project__code')
        )
        return Response(list(data))

    @action(detail=False, methods=['get'])
    def weekly(self, request):
        year = request.query_params.get('year', datetime.date.today().year)
        week = request.query_params.get('week', datetime.date.today().isocalendar()[1])
        entries = self.get_queryset().filter(year=year, week_number=week)
        serializer = self.get_serializer(entries, many=True)
        return Response(serializer.data)